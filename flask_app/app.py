"""
ISGA Gymnastics Scoring System
Flask Application - Compatible with PyCharm IDE
"""
import io
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, redirect, url_for, flash, request,
                   send_file, jsonify, abort)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate

from config import Config
from models import db, User, Event, Category, School, SchoolGroup, Athlete, Score
from forms import (LoginForm, RegistrationForm, EventForm, CategoryForm,
                   SchoolForm, AthleteForm, ScoreForm, AddSchoolToCategory)

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
migrate = Migrate(app, db)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


# ─── AUTH ROUTES ───────────────────────────────────────────────
@app.route('/')
def index():
    events = Event.query.order_by(Event.created_at.desc()).all()
    return render_template('index.html', events=events)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Logged in successfully.', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html', form=form)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RegistrationForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('Username already taken.', 'danger')
            return render_template('register.html', form=form)
        if User.query.filter_by(email=form.email.data).first():
            flash('Email already registered.', 'danger')
            return render_template('register.html', form=form)
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html', form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out.', 'info')
    return redirect(url_for('index'))


# ─── EVENT ROUTES ──────────────────────────────────────────────
@app.route('/event/create', methods=['GET', 'POST'])
@login_required
def create_event():
    form = EventForm()
    if form.validate_on_submit():
        event = Event(name=form.name.data, date=form.date.data)
        db.session.add(event)
        db.session.commit()
        flash('Event created.', 'success')
        return redirect(url_for('view_event', event_id=event.id))
    return render_template('create_event.html', form=form)


@app.route('/event/<int:event_id>')
def view_event(event_id):
    event = Event.query.get_or_404(event_id)
    return render_template('view_event.html', event=event)


@app.route('/event/<int:event_id>/add_category', methods=['GET', 'POST'])
@login_required
def add_category(event_id):
    event = Event.query.get_or_404(event_id)
    form = CategoryForm()
    if form.validate_on_submit():
        cat = Category(name=form.name.data, piece_type=form.piece_type.data, event_id=event.id)
        db.session.add(cat)
        db.session.commit()
        flash('Category added.', 'success')
        return redirect(url_for('view_event', event_id=event.id))
    return render_template('add_category.html', form=form, event=event)


# ─── SCHOOL ROUTES ─────────────────────────────────────────────
@app.route('/schools', methods=['GET', 'POST'])
@login_required
def manage_schools():
    form = SchoolForm()
    if form.validate_on_submit():
        if not School.query.filter_by(name=form.name.data).first():
            db.session.add(School(name=form.name.data))
            db.session.commit()
            flash('School added.', 'success')
        else:
            flash('School already exists.', 'warning')
    schools = School.query.order_by(School.name).all()
    return render_template('schools.html', form=form, schools=schools)


# ─── CATEGORY / SCORING ROUTES ─────────────────────────────────
@app.route('/category/<int:cat_id>')
def view_category(cat_id):
    category = Category.query.get_or_404(cat_id)
    # Recalculate all totals
    recalculate_category(category)
    return render_template('view_category.html', category=category)


@app.route('/category/<int:cat_id>/add_school', methods=['GET', 'POST'])
@login_required
def add_school_to_category(cat_id):
    category = Category.query.get_or_404(cat_id)
    form = AddSchoolToCategory()
    form.school_id.choices = [(s.id, s.name) for s in School.query.order_by(School.name).all()]
    if form.validate_on_submit():
        existing = SchoolGroup.query.filter_by(
            school_id=form.school_id.data, category_id=cat_id).first()
        if existing:
            flash('School already in this category.', 'warning')
        else:
            sg = SchoolGroup(school_id=form.school_id.data, category_id=cat_id)
            db.session.add(sg)
            db.session.flush()
            # Get max athlete number in this category
            max_num = db.session.query(db.func.max(Athlete.number)).join(SchoolGroup).filter(
                SchoolGroup.category_id == cat_id).scalar() or 0
            for i in range(form.num_athletes.data):
                max_num += 1
                athlete = Athlete(number=max_num, name='', school_group_id=sg.id)
                db.session.add(athlete)
                db.session.flush()
                db.session.add(Score(athlete_id=athlete.id))
            db.session.commit()
            flash('School and athletes added.', 'success')
        return redirect(url_for('view_category', cat_id=cat_id))
    return render_template('add_school_to_category.html', form=form, category=category)


@app.route('/category/<int:cat_id>/add_athlete/<int:sg_id>', methods=['POST'])
@login_required
def add_athlete_row(cat_id, sg_id):
    """Dynamically add a new athlete row."""
    sg = SchoolGroup.query.get_or_404(sg_id)
    max_num = db.session.query(db.func.max(Athlete.number)).join(SchoolGroup).filter(
        SchoolGroup.category_id == cat_id).scalar() or 0
    athlete = Athlete(number=max_num + 1, name='', school_group_id=sg.id)
    db.session.add(athlete)
    db.session.flush()
    db.session.add(Score(athlete_id=athlete.id))
    db.session.commit()
    flash('Athlete row added.', 'success')
    return redirect(url_for('score_input', cat_id=cat_id, sg_id=sg_id))


@app.route('/category/<int:cat_id>/score/<int:sg_id>', methods=['GET', 'POST'])
@login_required
def score_input(cat_id, sg_id):
    category = Category.query.get_or_404(cat_id)
    school_group = SchoolGroup.query.get_or_404(sg_id)
    athletes = Athlete.query.filter_by(school_group_id=sg_id).order_by(Athlete.number).all()

    if request.method == 'POST':
        for athlete in athletes:
            prefix = f'athlete_{athlete.id}_'
            athlete.name = request.form.get(f'{prefix}name', '')
            score = athlete.scores
            if not score:
                score = Score(athlete_id=athlete.id)
                db.session.add(score)
            score.set_vault = float(request.form.get(f'{prefix}set_vault', 0) or 0)
            score.vol_vault = float(request.form.get(f'{prefix}vol_vault', 0) or 0)
            score.set_floor = float(request.form.get(f'{prefix}set_floor', 0) or 0)
            score.vol_floor = float(request.form.get(f'{prefix}vol_floor', 0) or 0)
            score.individual_total = round(
                score.set_vault + score.vol_vault + score.set_floor + score.vol_floor, 2)
        db.session.commit()
        # Recalculate team totals
        recalculate_category(category)
        flash('Scores saved and totals recalculated.', 'success')
        return redirect(url_for('view_category', cat_id=cat_id))

    return render_template('score_input.html', category=category,
                           school_group=school_group, athletes=athletes)


# ─── AUTO-CALCULATION ──────────────────────────────────────────
def recalculate_category(category):
    """Recalculate team totals, positions, and group scores for entire category."""
    groups = SchoolGroup.query.filter_by(category_id=category.id).all()

    # Calculate individual positions across entire category
    all_athletes = []
    for g in groups:
        for a in g.athletes:
            if a.scores:
                a.scores.individual_total = round(
                    a.scores.set_vault + a.scores.vol_vault +
                    a.scores.set_floor + a.scores.vol_floor, 2)
                all_athletes.append(a)

    all_athletes.sort(key=lambda a: a.scores.individual_total, reverse=True)
    for pos, athlete in enumerate(all_athletes, 1):
        athlete.scores.individual_position = pos

    # Calculate team totals (top 4 scores per apparatus)
    for g in groups:
        scores_list = [a.scores for a in g.athletes if a.scores]
        top_n = 4  # top 4 scores per apparatus

        sv = sorted([s.set_vault for s in scores_list], reverse=True)[:top_n]
        vv = sorted([s.vol_vault for s in scores_list], reverse=True)[:top_n]
        sf = sorted([s.set_floor for s in scores_list], reverse=True)[:top_n]
        vf = sorted([s.vol_floor for s in scores_list], reverse=True)[:top_n]

        g.team_set_vault = round(sum(sv), 2)
        g.team_vol_vault = round(sum(vv), 2)
        g.team_set_floor = round(sum(sf), 2)
        g.team_vol_floor = round(sum(vf), 2)
        g.team_total = round(g.team_set_vault + g.team_vol_vault +
                             g.team_set_floor + g.team_vol_floor, 2)

    # Team positions
    groups.sort(key=lambda g: g.team_total, reverse=True)
    for pos, g in enumerate(groups, 1):
        g.team_position = pos

    # Group score = average of top 4 apparatus averages
    for g in groups:
        apparatus_avgs = []
        if g.team_set_vault > 0:
            apparatus_avgs.append(g.team_set_vault / 4)
        if g.team_vol_vault > 0:
            apparatus_avgs.append(g.team_vol_vault / 4)
        if g.team_set_floor > 0:
            apparatus_avgs.append(g.team_set_floor / 4)
        if g.team_vol_floor > 0:
            apparatus_avgs.append(g.team_vol_floor / 4)
        g.group_score = round(sum(apparatus_avgs), 2) if apparatus_avgs else 0

    # Group positions
    groups.sort(key=lambda g: g.group_score, reverse=True)
    for pos, g in enumerate(groups, 1):
        g.group_position = pos

    db.session.commit()


# ─── EXPORT ROUTES ─────────────────────────────────────────────
@app.route('/category/<int:cat_id>/export/excel')
@login_required
def export_excel(cat_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    category = Category.query.get_or_404(cat_id)
    recalculate_category(category)
    wb = Workbook()
    ws = wb.active
    ws.title = category.name

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6B2D7B', end_color='6B2D7B', fill_type='solid')
    cyan_fill = PatternFill(start_color='00BCD4', end_color='00BCD4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{category.event.name} - {category.name}"
    ws['A1'].font = Font(bold=True, size=14, color='6B2D7B')

    row = 3
    for sg in category.groups:
        # School header
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = sg.school.name
        ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='6B2D7B', end_color='6B2D7B', fill_type='solid')
        row += 1

        # Column headers
        headers = ['No', 'Name', 'Set Vault', 'Vol Vault', 'Set Floor', 'Vol Floor', 'Total', 'Pos']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = cyan_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # Team totals headers
        ws.cell(row=row, column=10, value='Team Totals').font = Font(bold=True)
        ws.cell(row=row, column=14, value='Score').font = Font(bold=True)
        ws.cell(row=row, column=15, value='Pos.').font = Font(bold=True)
        row += 1

        # Athletes
        team_row_start = row
        for i, athlete in enumerate(sg.athletes):
            s = athlete.scores
            ws.cell(row=row, column=1, value=athlete.number).border = thin_border
            ws.cell(row=row, column=2, value=athlete.name).border = thin_border
            ws.cell(row=row, column=3, value=s.set_vault if s else 0).border = thin_border
            ws.cell(row=row, column=4, value=s.vol_vault if s else 0).border = thin_border
            ws.cell(row=row, column=5, value=s.set_floor if s else 0).border = thin_border
            ws.cell(row=row, column=6, value=s.vol_floor if s else 0).border = thin_border
            ws.cell(row=row, column=7, value=s.individual_total if s else 0).border = thin_border
            ws.cell(row=row, column=8, value=s.individual_position if s else 0).border = thin_border

            # Team totals column (first 4 rows)
            if i == 0:
                ws.cell(row=row, column=10, value='Set Vault')
                ws.cell(row=row, column=11, value=sg.team_set_vault)
            elif i == 1:
                ws.cell(row=row, column=10, value='Vol Vault')
                ws.cell(row=row, column=11, value=sg.team_vol_vault)
            elif i == 2:
                ws.cell(row=row, column=10, value='Set Floor')
                ws.cell(row=row, column=11, value=sg.team_set_floor)
            elif i == 3:
                ws.cell(row=row, column=10, value='Vol Floor')
                ws.cell(row=row, column=11, value=sg.team_vol_floor)

            # Group score on first row
            if i == 0:
                ws.cell(row=row, column=13, value='Group')
                ws.cell(row=row, column=14, value=sg.group_score)
                ws.cell(row=row, column=15, value=sg.group_position)
            row += 1

        # Team total row
        ws.cell(row=row, column=2, value='Team Totals').font = Font(bold=True)
        ws.cell(row=row, column=3, value=sg.team_set_vault).font = Font(bold=True)
        ws.cell(row=row, column=4, value=sg.team_vol_vault).font = Font(bold=True)
        ws.cell(row=row, column=5, value=sg.team_set_floor).font = Font(bold=True)
        ws.cell(row=row, column=6, value=sg.team_vol_floor).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=8, value=sg.team_total).font = Font(bold=True, size=12)
        ws.cell(row=row, column=9, value=sg.team_position).font = Font(bold=True)
        row += 2

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{category.event.name}_{category.name}.xlsx".replace(' ', '_')
    return send_file(output, download_name=filename,
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/category/<int:cat_id>/export/pdf')
@login_required
def export_pdf(cat_id):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    category = Category.query.get_or_404(cat_id)
    recalculate_category(category)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15 * mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                 textColor=colors.HexColor('#6B2D7B'), fontSize=16)
    elements.append(Paragraph(f"{category.event.name} - {category.name}", title_style))
    elements.append(Spacer(1, 10 * mm))

    isga_purple = colors.HexColor('#6B2D7B')
    isga_cyan = colors.HexColor('#00BCD4')

    for sg in category.groups:
        elements.append(Paragraph(f"<b>{sg.school.name}</b>", styles['Heading3']))
        data = [['No', 'Name', 'Set Vault', 'Vol Vault', 'Set Floor', 'Vol Floor',
                 'Total', 'Pos', '', 'Team Totals', '', '', 'Group', 'Score', 'Pos.']]

        for i, athlete in enumerate(sg.athletes):
            s = athlete.scores
            row = [athlete.number, athlete.name,
                   f'{s.set_vault:.2f}' if s else '0.00',
                   f'{s.vol_vault:.2f}' if s else '0.00',
                   f'{s.set_floor:.2f}' if s else '0.00',
                   f'{s.vol_floor:.2f}' if s else '0.00',
                   f'{s.individual_total:.2f}' if s else '0.00',
                   s.individual_position if s else 0, '']
            if i == 0:
                row += ['Set Vault', f'{sg.team_set_vault:.2f}', '',
                        'Group', f'{sg.group_score:.2f}', sg.group_position]
            elif i == 1:
                row += ['Vol Vault', f'{sg.team_vol_vault:.2f}', '', '', '', '']
            elif i == 2:
                row += ['Set Floor', f'{sg.team_set_floor:.2f}', '', '', '', '']
            elif i == 3:
                row += ['Vol Floor', f'{sg.team_vol_floor:.2f}', '', '', '', '']
            else:
                row += ['', '', '', '', '', '']
            data.append(row)

        # Team totals row
        data.append(['', 'Team Totals',
                     f'{sg.team_set_vault:.2f}', f'{sg.team_vol_vault:.2f}',
                     f'{sg.team_set_floor:.2f}', f'{sg.team_vol_floor:.2f}',
                     '', '', '', f'Total: {sg.team_total:.2f}', f'Pos: {sg.team_position}',
                     '', '', '', ''])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), isga_cyan),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (7, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (7, -1), colors.HexColor('#F0E6F6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (7, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 8 * mm))

    doc.build(elements)
    output.seek(0)
    filename = f"{category.event.name}_{category.name}.pdf".replace(' ', '_')
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/pdf')


# ─── INIT DB & CREATE ADMIN ───────────────────────────────────
@app.cli.command('init-db')
def init_db():
    """Initialize database and create admin user."""
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', email='admin@isga.org', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print('Admin user created: admin / admin123')
    print('Database initialized.')


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        #if not User.query.filter_by(username='admin').first():
         #   admin = User(username='admin', email='admin@isga.org', role='admin')
          #  admin.set_password('admin123')
           # db.session.add(admin)
            #db.session.commit()
    app.run(debug=True, port=5000)
